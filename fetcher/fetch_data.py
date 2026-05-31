"""
WECC Hydro Market Brief — Daily Data Fetcher
─────────────────────────────────────────────
Runs daily at 06:00 PT via GitHub Actions. Pulls from primary public
sources, writes a single JSON file consumed by the static dashboard.

Design rules
────────────
1. Graceful degradation: any source can fail without killing the run.
2. Primary sources only. No paywalled feeds. No scraping behind logins.
3. One JSON output. Frontend reads /data/dashboard.json — that's it.
4. Idempotent. Safe to re-run any time; output always reflects "now".

Environment
───────────
    EIA_API_KEY    Required for BPA mix.
                   Free at https://www.eia.gov/opendata/register.php
                   Set as GitHub Actions secret.
"""

import os
import sys
import json
import re
import datetime
import urllib.parse
from pathlib import Path
from typing import Optional, Dict, List, Any
from xml.etree import ElementTree as ET

import requests

# ── Config ────────────────────────────────────────────────────────────
EIA_API_KEY  = os.environ.get("EIA_API_KEY", "")
OUT_PATH     = Path(__file__).resolve().parent.parent / "data" / "dashboard.json"
ARCHIVE_PATH = Path(__file__).resolve().parent.parent / "data" / "archive.json"
PT           = datetime.timezone(datetime.timedelta(hours=-7))
HTTP_TIMEOUT = 30
UA           = "WECC-Hydro-Brief/1.0 (https://wecchydrobrief.com)"

# Basin roster — 17 basins, north-to-south. Names MUST match the
# `name` property in data/geo/basins.geojson so the map renders.
BASINS: List[Dict[str, str]] = [
    # BC Columbia headwaters
    {"name": "Mica / Upper Columbia (BC)", "source": "bc",     "id": "UCOL"},
    {"name": "Kootenay (BC)",              "source": "bc",     "id": "KOOT"},
    # PNW (Columbia + Snake)
    {"name": "Upper Columbia (US)",        "source": "snotel", "id": "170200"},
    {"name": "Pend Oreille",               "source": "snotel", "id": "170102"},
    {"name": "Yakima",                     "source": "snotel", "id": "170300"},
    {"name": "Clearwater",                 "source": "snotel", "id": "170603"},
    {"name": "Salmon",                     "source": "snotel", "id": "170602"},
    {"name": "Upper Snake",                "source": "snotel", "id": "170401"},
    {"name": "Lower Snake",                "source": "snotel", "id": "170601"},
    {"name": "Owyhee",                     "source": "snotel", "id": "170501"},
    {"name": "Lower Columbia",             "source": "snotel", "id": "170800"},
    # CAISO Sierra
    {"name": "Northern Sierra (Feather)",  "source": "cdec",   "id": "NSF"},
    {"name": "Central Sierra (American)",  "source": "cdec",   "id": "CSF"},
    {"name": "Southern Sierra (San Joaq.)","source": "cdec",   "id": "SSF"},
    {"name": "Tulare (Kings/Kern)",        "source": "cdec",   "id": "TLR"},
    # Colorado (no SWE source feed; basin shows on map but stays "no data")
    {"name": "Upper Colorado",             "source": "none",   "id": "UCOLO"},
    {"name": "Lower Colorado",             "source": "none",   "id": "LCOLO"},
]

# USACE NWD projects to fetch discharge for.
USACE_PROJECTS: List[Dict[str, str]] = [
    {"name": "Grand Coulee", "code": "GCL"},
    {"name": "Chief Joseph", "code": "CHJ"},
    {"name": "Lower Granite","code": "LWG"},
    {"name": "The Dalles",   "code": "TDA"},
    {"name": "Bonneville",   "code": "BON"},
]


# ── Utility ───────────────────────────────────────────────────────────
def _log(msg: str) -> None:
    print(msg, flush=True)


def _safe(fn, label: str):
    try:
        return fn()
    except Exception as e:
        _log(f"  ✗ {label}: {type(e).__name__}: {e}")
        return None


def _http_get(url: str, **kwargs) -> requests.Response:
    headers = kwargs.pop("headers", {})
    headers.setdefault("User-Agent", UA)
    r = requests.get(url, headers=headers, timeout=HTTP_TIMEOUT, **kwargs)
    r.raise_for_status()
    return r


# ══════════════════════════════════════════════════════════════════════
#  TIER 1 — EIA-930 (working) + EIA ICE bulk XLS (new)
# ══════════════════════════════════════════════════════════════════════

def fetch_eia_bpa_mix() -> Optional[Dict[str, int]]:
    """
    EIA-930 BPA balancing-authority hourly generation by fuel type.
    Averages the last 24h into hydro / wind / thermal-and-other.
    """
    if not EIA_API_KEY:
        _log("  ⚠ EIA_API_KEY unset — skipping BPA mix")
        return None

    end   = datetime.datetime.utcnow()
    start = end - datetime.timedelta(days=2)
    url   = "https://api.eia.gov/v2/electricity/rto/fuel-type-data/data/"
    params = {
        "api_key": EIA_API_KEY,
        "frequency": "hourly",
        "data[0]": "value",
        "facets[respondent][]": "BPAT",
        "start": start.strftime("%Y-%m-%dT%H"),
        "end":   end.strftime("%Y-%m-%dT%H"),
        "sort[0][column]": "period",
        "sort[0][direction]": "desc",
        "length": 5000,
    }
    rows = _http_get(url, params=params).json().get("response", {}).get("data", [])
    if not rows:
        return None

    by_hour: Dict[str, Dict[str, float]] = {}
    for row in rows:
        h, ft, v = row.get("period"), row.get("fueltype"), row.get("value")
        if h is None or ft is None or v is None:
            continue
        by_hour.setdefault(h, {})[ft] = float(v)
    last_24 = sorted(by_hour.keys(), reverse=True)[:24]

    totals: Dict[str, float] = {}
    for h in last_24:
        for ft, v in by_hour[h].items():
            totals[ft] = totals.get(ft, 0.0) + v

    grand = sum(totals.values()) or 1.0
    hydro = totals.get("WAT", 0.0) / grand * 100
    wind  = totals.get("WND", 0.0) / grand * 100
    therm = 100.0 - hydro - wind
    _log(f"  ✓ BPA mix: hydro {hydro:.0f}% / wind {wind:.0f}% / other {therm:.0f}%")
    return {
        "hydro_pct":   round(hydro),
        "wind_pct":    round(wind),
        "thermal_pct": round(therm),
    }


def fetch_eia_midc_price() -> Optional[Dict[str, Any]]:
    """
    Download EIA's bulk ICE wholesale electricity XLS, parse Mid-C peak.

    Source: https://www.eia.gov/electricity/wholesale/
            The page links to ice_electric-YYYY-YYYY.xlsx (current year file).
    Why this and not Today in Energy: bulk XLS is an authoritative republication
    of ICE's daily indices under EIA's licensing agreement, includes peak +
    off-peak + weighted-avg + volume, history back to 2001. Today in Energy
    only shows the previous day's headline.
    """
    import openpyxl

    year = datetime.date.today().year
    # The file is named for the year range it covers. Current convention:
    # "ice_electric-2010-{year}.xlsx" — EIA expands the range each year.
    candidates = [
        f"https://www.eia.gov/electricity/wholesale/xls/ice_electric-2010-{year}.xlsx",
        f"https://www.eia.gov/electricity/wholesale/xls/ice_electric-2010-{year - 1}.xlsx",
    ]
    xls_bytes = None
    for url in candidates:
        try:
            r = _http_get(url)
            xls_bytes = r.content
            _log(f"  · fetched {url.rsplit('/', 1)[-1]} ({len(xls_bytes)//1024} KB)")
            break
        except requests.HTTPError:
            continue
    if xls_bytes is None:
        raise RuntimeError("EIA ICE bulk XLS not found at expected URLs")

    import io
    wb = openpyxl.load_workbook(io.BytesIO(xls_bytes), data_only=True, read_only=True)

    # Sheet naming varies by year; find sheets whose name contains "Mid C"
    # (the actual hub name in EIA's file is "Mid C Peak" / "Mid C Off-Peak").
    peak_sheet = None
    offp_sheet = None
    for name in wb.sheetnames:
        norm = name.lower().replace("-", " ").replace("  ", " ")
        if "mid c" in norm and "peak" in norm and "off" not in norm:
            peak_sheet = name
        elif "mid c" in norm and "off" in norm:
            offp_sheet = name
    if not peak_sheet:
        raise RuntimeError(f"No Mid-C peak sheet in {wb.sheetnames}")

    def read_sheet(name):
        """Returns list of (date, wtd_avg_price) tuples, newest last."""
        ws = wb[name]
        # Header row: identify columns. EIA layout: Trade Date, Delivery Start,
        # Delivery End, High, Low, Wtd Avg Price $/MWh, Daily Volume MWh, # Trades.
        header = None
        out = []
        for row in ws.iter_rows(values_only=True):
            if header is None:
                if row and isinstance(row[0], str) and "date" in row[0].lower():
                    header = [str(c or "").lower().strip() for c in row]
                continue
            if not row or row[0] is None:
                continue
            d = row[0]
            if isinstance(d, datetime.datetime):
                d = d.date()
            elif isinstance(d, str):
                try:
                    d = datetime.date.fromisoformat(d[:10])
                except ValueError:
                    continue
            # find wtd avg column
            try:
                wtd_idx = next(i for i, h in enumerate(header) if "wtd" in h and "avg" in h)
            except StopIteration:
                wtd_idx = 5  # fallback: typical EIA layout
            try:
                p = float(row[wtd_idx])
            except (TypeError, ValueError):
                continue
            out.append((d, p))
        return out

    peak_series = read_sheet(peak_sheet)
    offp_series = read_sheet(offp_sheet) if offp_sheet else []

    if not peak_series:
        raise RuntimeError("Peak series empty")

    peak_series.sort()
    if offp_series:
        offp_series.sort()

    latest_date, latest_price = peak_series[-1]
    offp_latest = offp_series[-1][1] if offp_series else None

    # 7-day WoW
    week_ago = latest_date - datetime.timedelta(days=7)
    prior = next((p for d, p in reversed(peak_series[:-1]) if d <= week_ago), None)
    wow_pct = ((latest_price - prior) / prior * 100) if prior else None

    # 30-day chart history
    history = [
        {"d": _fmt_short_date(d.isoformat()), "p": round(p, 2)}
        for d, p in peak_series[-30:]
    ]

    _log(f"  ✓ Mid-C peak DA: ${latest_price:.2f}/MWh ({latest_date.isoformat()})")
    return {
        "peak_da":    round(latest_price, 2),
        "offpeak_da": round(offp_latest, 2) if offp_latest is not None else None,
        "wow_pct":    round(wow_pct, 1) if wow_pct is not None else None,
        "history":    history,
        "as_of":      latest_date.isoformat(),
    }


def _fmt_short_date(iso: str) -> str:
    d = datetime.date.fromisoformat(iso[:10])
    return d.strftime("%b %-d") if sys.platform != "win32" else d.strftime("%b %#d")


# ══════════════════════════════════════════════════════════════════════
#  TIER 2 STUBS (filled out in future sessions)
# ══════════════════════════════════════════════════════════════════════

def fetch_snotel_basin(huc: str) -> Optional[Dict[str, Any]]:
    """
    NRCS AWDB REST API v2 — basin-averaged SWE % of median for a HUC-6.

    Strategy:
    1. Query all active SNOTEL stations in the HUC-6.
    2. For each station get today's WTEQ (snow water equiv, inches) and
       the 1991-2020 median for the same calendar date.
    3. Return basin-weighted average pct_median, mean swe_in, and 7-day delta.

    Source: https://wcc.sc.egov.usda.gov/awdbRestApi/swagger-ui/index.html
    """
    today = datetime.date.today()
    week_ago = today - datetime.timedelta(days=7)
    base = "https://wcc.sc.egov.usda.gov/awdbRestApi/services/v1"

    # Step 1: get station list for this HUC-6
    station_url = f"{base}/stations"
    params = {
        "hucs": huc,
        "networkCds": "SNTL",
        "activeStationsOnly": "true",
    }
    stations = _http_get(station_url, params=params).json()
    if not stations:
        _log(f"    ⚠ SNOTEL {huc}: no stations found")
        return None
    triplets = [s["stationTriplet"] for s in stations]

    # Step 2: fetch current WTEQ + 7-day-ago WTEQ for all stations
    data_url = f"{base}/data"
    def get_wteq(date_str):
        p = {
            "stationTriplets": ",".join(triplets),
            "elementCd": "WTEQ",
            "beginDate": date_str,
            "endDate": date_str,
            "duration": "DAILY",
            "getFlags": "false",
        }
        rows = _http_get(data_url, params=p).json()
        vals = {}
        for row in rows:
            v = row.get("values", [{}])
            val = v[0].get("value") if v else None
            if val is not None:
                try:
                    vals[row["stationTriplet"]] = float(val)
                except (TypeError, ValueError):
                    pass
        return vals

    today_wteq = get_wteq(today.isoformat())
    week_wteq  = get_wteq(week_ago.isoformat())

    if not today_wteq:
        _log(f"    ⚠ SNOTEL {huc}: no WTEQ values returned")
        return None

    # Step 3: fetch 1991-2020 median for today's date
    normals_url = f"{base}/stationElements"
    valid_triplets = list(today_wteq.keys())
    median_vals = {}
    # Normals endpoint works per-station; batch with comma-separated triplets
    norm_params = {
        "stationTriplets": ",".join(valid_triplets),
        "elementCd": "WTEQ",
        "periodRef": "CURRENT",
        "beginMonthDay": today.strftime("%m-%d"),
        "endMonthDay": today.strftime("%m-%d"),
        "duration": "DAILY",
    }
    try:
        norm_rows = _http_get(normals_url, params=norm_params).json()
        for row in norm_rows:
            med = row.get("normalValue")
            if med is not None:
                try:
                    median_vals[row["stationTriplet"]] = float(med)
                except (TypeError, ValueError):
                    pass
    except Exception:
        pass  # normals are best-effort; pct_median degrades gracefully

    # Step 4: compute averages
    swe_vals = list(today_wteq.values())
    mean_swe = sum(swe_vals) / len(swe_vals)

    # pct_median: average (obs/median)*100 across stations that have both
    pct_list = []
    for t, obs in today_wteq.items():
        med = median_vals.get(t)
        if med and med > 0:
            pct_list.append(obs / med * 100)
    pct_median = round(sum(pct_list) / len(pct_list)) if pct_list else None

    # 7-day delta (% of median units — same calculation on week-ago values)
    if week_wteq:
        week_pct_list = []
        for t, obs in week_wteq.items():
            med = median_vals.get(t)
            if med and med > 0:
                week_pct_list.append(obs / med * 100)
        if week_pct_list and pct_list:
            week_pct = sum(week_pct_list) / len(week_pct_list)
            delta_7d = round((sum(pct_list) / len(pct_list)) - week_pct, 1)
        else:
            delta_7d = None
    else:
        delta_7d = None

    return {
        "pct_median": pct_median,
        "swe_in":     round(mean_swe, 1),
        "delta_7d":   delta_7d,
        "n_stations": len(swe_vals),
        "as_of":      today.isoformat(),
    }


def fetch_cdec_sierra(region_id: str) -> Optional[Dict[str, Any]]:
    """
    CDEC (CA Dept of Water Resources) regional snow sensor data.
    Uses the CDEC Water Data Library CSV export for regional SWE averages.

    region_id maps to CDEC station group IDs:
      NSF = North Fork Feather / Northern Sierra
      CSF = Central Sierra (American)
      SSF = Southern Sierra (San Joaquin)
      TLR = Tulare (Kings/Kern)

    CDEC publishes regional % of average SWE via:
    https://cdec.water.ca.gov/cgi-progs/products/DLYSWE.html  (HTML, complex)

    Alternative: CDEC snow sensor CSV endpoint for individual stations.
    We use the CDO-style regional summaries at:
    https://cdec.water.ca.gov/snow/current/snow/index.html

    Practical note: CDEC's regional SWE % of average is also published
    in a clean CSV at:
    https://cdec.water.ca.gov/reportapp/javareports?name=PLOT_SWE
    which returns HTML with embedded tables.

    Cleanest machine-readable source: CDEC WSPF snow water content CSV:
    https://cdec.water.ca.gov/cgi-progs/products/swcchart.csv
    This returns a CSV with columns: Region, Current SWE (in), Average SWE (in), % of Average
    Regions: NW, CN, SW (North, Central, South Sierra) + state total.
    """
    region_map = {
        "NSF": "NW",
        "CSF": "CN",
        "SSF": "SW",
        "TLR": "SW",  # Tulare falls in Southern Sierra aggregate
    }
    cdec_region = region_map.get(region_id)
    if not cdec_region:
        return None

    url = "https://cdec.water.ca.gov/cgi-progs/products/swcchart.csv"
    r = _http_get(url)
    today = datetime.date.today()

    # Parse CSV: skip comment lines, find header, then data rows
    lines = [ln.strip() for ln in r.text.splitlines() if ln.strip()]
    header_idx = None
    for i, ln in enumerate(lines):
        if "Region" in ln or "region" in ln.lower():
            header_idx = i
            break
    if header_idx is None:
        # Fallback: try parsing as fixed-width or comma-separated with positional logic
        # Look for lines containing the region code
        for ln in lines:
            parts = [p.strip() for p in ln.split(",")]
            if len(parts) >= 4 and parts[0].upper() == cdec_region:
                try:
                    swe_cur = float(parts[1])
                    swe_avg = float(parts[2])
                    pct = round(swe_cur / swe_avg * 100) if swe_avg else None
                    return {
                        "pct_median": pct,
                        "swe_in":     round(swe_cur, 1),
                        "delta_7d":   None,
                        "n_stations": None,
                        "as_of":      today.isoformat(),
                    }
                except (ValueError, ZeroDivisionError):
                    pass
        return None

    headers = [h.strip().lower() for h in lines[header_idx].split(",")]
    for ln in lines[header_idx + 1:]:
        parts = [p.strip() for p in ln.split(",")]
        if not parts or parts[0].upper() != cdec_region:
            continue
        try:
            row = dict(zip(headers, parts))
            # Column names vary by year; try common variants
            cur_key  = next((k for k in row if "current" in k or "swe" in k and "avg" not in k), None)
            avg_key  = next((k for k in row if "average" in k or "avg" in k), None)
            pct_key  = next((k for k in row if "%" in k or "percent" in k), None)
            swe_cur  = float(row[cur_key]) if cur_key else None
            swe_avg  = float(row[avg_key]) if avg_key else None
            pct_str  = row.get(pct_key, "").replace("%", "").strip()
            pct      = int(float(pct_str)) if pct_str else (
                       round(swe_cur / swe_avg * 100) if swe_cur and swe_avg and swe_avg > 0 else None)
            return {
                "pct_median": pct,
                "swe_in":     round(swe_cur, 1) if swe_cur is not None else None,
                "delta_7d":   None,
                "n_stations": None,
                "as_of":      today.isoformat(),
            }
        except (ValueError, TypeError, KeyError):
            continue
    return None


def fetch_bc_snow(region_id: str) -> Optional[Dict[str, Any]]:
    """
    BC River Forecast Centre automated snow weather station data.
    BC RFC publishes a bulletin PDF (hard to parse) and an interactive map.
    The cleanest machine-readable source is the BC RFC Snow Conditions CSV:
    https://www.env.gov.bc.ca/wsd/data_searches/snow/asws/data/

    For % of normal, we use the BC RFC Water Supply Bulletin data table
    served from their public API. As a reliable fallback, we return None
    and let the map show "no data" for BC basins (still renders correctly).

    Implemented: query ASWS (Automated Snow Weather Stations) via:
    https://www.env.gov.bc.ca/wsd/data_searches/snow/asws/data/current_conditions.csv
    which has columns: Station, Basin, SWE(mm), %Normal, etc.
    """
    basin_map = {
        "UCOL": ["Upper Columbia"],
        "KOOT": ["Kootenay"],
    }
    target_basins = basin_map.get(region_id, [])
    if not target_basins:
        return None

    url = "https://www.env.gov.bc.ca/wsd/data_searches/snow/asws/data/current_conditions.csv"
    try:
        r = _http_get(url)
    except Exception:
        return None

    lines = [ln.strip() for ln in r.text.splitlines() if ln.strip()]
    if len(lines) < 2:
        return None

    headers = [h.strip().lower().replace(" ", "_") for h in lines[0].split(",")]
    basin_col  = next((i for i, h in enumerate(headers) if "basin" in h), None)
    swe_col    = next((i for i, h in enumerate(headers) if "swe" in h and "%" not in h), None)
    pct_col    = next((i for i, h in enumerate(headers) if "%" in h or "normal" in h or "median" in h), None)
    if basin_col is None or pct_col is None:
        return None

    pct_list = []
    swe_list = []
    for ln in lines[1:]:
        parts = [p.strip() for p in ln.split(",")]
        if len(parts) <= max(filter(None, [basin_col, pct_col, swe_col])):
            continue
        basin_val = parts[basin_col].strip()
        if not any(tb.lower() in basin_val.lower() for tb in target_basins):
            continue
        try:
            pct = float(parts[pct_col].replace("%", ""))
            pct_list.append(pct)
        except (ValueError, IndexError):
            pass
        if swe_col is not None:
            try:
                swe_mm = float(parts[swe_col])
                swe_list.append(swe_mm * 0.0394)  # mm → inches
            except (ValueError, IndexError):
                pass

    if not pct_list:
        return None

    today = datetime.date.today()
    return {
        "pct_median": round(sum(pct_list) / len(pct_list)),
        "swe_in":     round(sum(swe_list) / len(swe_list), 1) if swe_list else None,
        "delta_7d":   None,
        "n_stations": len(pct_list),
        "as_of":      today.isoformat(),
    }

def fetch_nwrfc_wsf() -> Optional[Dict[str, Any]]:
    """
    NWRFC Apr–Sep unregulated water supply forecast at The Dalles.

    Source: NWRFC publishes a CSV table of current-year volume forecasts
    at https://www.nwrfc.noaa.gov/water_supply/ws_text.php?id=TDAO3&wfo=
    That page returns a plain-text table. We parse the most recent
    forecast and the 1981-2010 normal to compute % of normal.

    Fallback: NWRFC also publishes ensemble data at:
    https://www.nwrfc.noaa.gov/ws/ws_api.php  (GeoJSON-style JSON)
    """
    # Primary: NWRFC Water Supply text report for The Dalles (TDAO3)
    url = "https://www.nwrfc.noaa.gov/water_supply/ws_text.php"
    params = {"id": "TDAO3", "wfo": ""}
    try:
        r = _http_get(url, params=params)
        text = r.text
    except Exception as e:
        _log(f"  ✗ NWRFC WSF primary: {e}")
        return None

    # The text report includes lines like:
    # "Apr-Sep  1234  MAF  123%  ..."
    # We scan for a line with Apr-Sep and extract volume + percent
    pct_normal = None
    forecast_maf = None

    for line in text.splitlines():
        line = line.strip()
        # Look for lines mentioning Apr-Sep and a percentage
        if re.search(r"apr.?sep", line, re.IGNORECASE):
            # Extract percentage
            m_pct = re.search(r"(\d{1,3})\s*%", line)
            if m_pct:
                pct_normal = int(m_pct.group(1))
            # Extract MAF (million acre-feet): number before "MAF" or after "%" in a numeric field
            m_maf = re.search(r"(\d+\.?\d*)\s*(?:MAF|maf)", line)
            if m_maf:
                forecast_maf = float(m_maf.group(1))
            if pct_normal is not None:
                break

    # Fallback: scan for any percentage on a line with the forecast period
    if pct_normal is None:
        for line in text.splitlines():
            m = re.search(r"(\d{2,4})\s+(\d{1,3})%", line)
            if m:
                forecast_maf = float(m.group(1))
                pct_normal   = int(m.group(2))
                break

    if pct_normal is None:
        _log("  ✗ NWRFC WSF: could not parse percent of normal from response")
        return None

    _log(f"  ✓ NWRFC WSF: {pct_normal}% of normal ({forecast_maf} MAF)")
    return {
        "site":         "The Dalles",
        "pct_normal":   pct_normal,
        "forecast_maf": forecast_maf,
        "as_of":        datetime.date.today().isoformat(),
    }


def fetch_usace_project(code: str) -> Optional[Dict[str, Any]]:
    """
    USACE NWD dataquery — current discharge (kcfs) and forebay elevation (ft).

    Source: USACE Northwestern Division Dataquery 2.0
    https://www.nwd-wc.usace.army.mil/dd/common/dataquery/www/

    The endpoint accepts project codes (e.g. GCL, BON) and returns JSON
    with the most recent observed values for outflow and forebay.
    """
    base = "https://www.nwd-wc.usace.army.mil/dd/common/dataquery/www/"
    # Parameters for daily outflow (QD) and forebay (FB) for today
    today = datetime.date.today()
    yesterday = today - datetime.timedelta(days=1)
    params = {
        "id":     code,
        "dt":     "daily",
        "variable": "QD,FB",
        "startdate": yesterday.strftime("%Y-%m-%d"),
        "enddate":   today.strftime("%Y-%m-%d"),
        "format":   "json",
    }
    try:
        r = _http_get(base, params=params)
        data = r.json()
    except Exception as e:
        _log(f"  ✗ USACE {code}: {e}")
        return None

    # Response structure: {"data": [{...},...], "sites": {...}}
    # Each data item has "site_id", "parameter", "values": [[timestamp, value], ...]
    discharge_kcfs = None
    forebay_ft     = None

    datasets = data if isinstance(data, list) else data.get("data", [])
    for ds in datasets:
        param  = str(ds.get("parameter", "")).upper()
        values = ds.get("values", [])
        if not values:
            continue
        # Take most recent non-null value
        for ts, val in reversed(values):
            if val is not None:
                try:
                    v = float(val)
                except (TypeError, ValueError):
                    continue
                if "QD" in param or "FLOW" in param or "OUTFLOW" in param:
                    discharge_kcfs = round(v, 1)
                elif "FB" in param or "FOREBAY" in param or "ELEV" in param:
                    forebay_ft = round(v, 1)
                break

    if discharge_kcfs is None and forebay_ft is None:
        _log(f"  ✗ USACE {code}: no usable data in response")
        return None

    _log(f"  ✓ USACE {code}: {discharge_kcfs} kcfs / {forebay_ft} ft")
    return {
        "discharge_kcfs": discharge_kcfs,
        "forebay_ft":     forebay_ft,
        "as_of":          today.isoformat(),
    }


def fetch_caiso_hydro() -> Optional[Dict[str, Any]]:
    """
    CAISO OASIS — current hydro dispatch (MW) and SP15 day-ahead price.

    Source: CAISO OASIS API (public, no auth required)
    Endpoint: http://oasis.caiso.com/oasisapi/SingleZip
    Query type: SLD_FCST (system load + renewable forecast, includes hydro)
    and PRC_LMP for SP15 prices.

    We use the ENE_SLRS (Renewables + Storage supply) report which includes
    hydro as a fuel type under "renewable" generation.

    Practical note: CAISO OASIS can be slow. We use the "AT" (Actual Trades)
    market results for hydro generation via the PROD_MW endpoint.
    """
    today     = datetime.date.today()
    yesterday = today - datetime.timedelta(days=1)
    base      = "http://oasis.caiso.com/oasisapi/SingleZip"

    def oasis_get(query_name, startdt, enddt, extra=None):
        params = {
            "queryname": query_name,
            "startdatetime": startdt.strftime("%Y%m%dT07:00-0000"),
            "enddatetime":   enddt.strftime("%Y%m%dT07:00-0000"),
            "version":       1,
        }
        if extra:
            params.update(extra)
        r = _http_get(base, params=params)
        import zipfile, io
        z = zipfile.ZipFile(io.BytesIO(r.content))
        # Find the XML file inside the zip
        xml_name = next((n for n in z.namelist() if n.endswith(".xml")), None)
        if not xml_name:
            return None
        return ET.fromstring(z.read(xml_name))

    hydro_mw    = None
    sp15_da     = None
    hydro_share = None
    history     = []

    # ── Hydro generation: SLD_REN_FCST (renewable + hydro supply) ──
    try:
        root = oasis_get("ENE_SLRS", yesterday, today,
                         extra={"market_run_id": "ACTUAL", "tac_zone_name": "ALL"})
        if root is not None:
            ns  = {"c": "http://www.caiso.com/soa/OASISReport_v1.xsd"}
            mw_vals = []
            for rpt in root.findall(".//c:REPORT_DATA", ns):
                fuel = rpt.findtext("c:FUEL_TYPE", namespaces=ns) or ""
                mkt  = rpt.findtext("c:MARKET_RUN_ID", namespaces=ns) or ""
                if "HYDR" in fuel.upper() and "ACTUAL" in mkt.upper():
                    val = rpt.findtext("c:MW", namespaces=ns)
                    if val:
                        try:
                            mw_vals.append(float(val))
                        except ValueError:
                            pass
            if mw_vals:
                hydro_mw = round(sum(mw_vals) / len(mw_vals))
    except Exception as e:
        _log(f"  · CAISO hydro MW: {type(e).__name__}: {e}")

    # ── SP15 DA LMP (day-ahead locational marginal price) ──
    try:
        root = oasis_get("PRC_LMP", yesterday, today,
                         extra={"market_run_id": "DAM", "node": "TH_SP15_GEN-APND"})
        if root is not None:
            ns = {"c": "http://www.caiso.com/soa/OASISReport_v1.xsd"}
            lmp_vals = []
            for rpt in root.findall(".//c:REPORT_DATA", ns):
                lmp_type = rpt.findtext("c:LMP_TYPE", namespaces=ns) or ""
                if lmp_type.upper() == "LMP":
                    val = rpt.findtext("c:MW", namespaces=ns)
                    if val:
                        try:
                            lmp_vals.append(float(val))
                        except ValueError:
                            pass
            if lmp_vals:
                sp15_da = round(sum(lmp_vals) / len(lmp_vals), 2)
    except Exception as e:
        _log(f"  · CAISO SP15 DA: {type(e).__name__}: {e}")

    if hydro_mw is None and sp15_da is None:
        _log("  ✗ CAISO: no data retrieved")
        return None

    _log(f"  ✓ CAISO: hydro {hydro_mw} MW / SP15 DA ${sp15_da}")
    return {
        "hydro_mw":        hydro_mw,
        "hydro_share_pct": hydro_share,
        "sp15_da":         sp15_da,
        "history":         history,
        "as_of":           today.isoformat(),
    }


# ══════════════════════════════════════════════════════════════════════
#  REGULATORY PULSE — FERC + CAISO + BPA (3 feeds, deduped, max 6 items)
# ══════════════════════════════════════════════════════════════════════

# Hydropower keywords for filtering noisy feeds. Matches case-insensitive.
HYDRO_KEYWORDS = re.compile(
    r"\b(hydro|hydropower|hydroelectric|pumped storage|psh|"
    r"dam|reservoir|spillway|spill|columbia|snake|"
    r"bpa|bonneville|edam|wem|weim|markets\+?|markets plus|"
    r"ferc license|relicens|water power|"
    r"colorado river|hoover|glen canyon|"
    r"caiso|cal[ -]?iso|wecc)\b",
    re.IGNORECASE
)


def fetch_ferc_filings(max_items: int = 5) -> List[Dict[str, str]]:
    """
    FERC eForms RSS — filter for hydropower-relevant filings.
    Source: https://ecollection.ferc.gov/api/rssfeed
    """
    url = "https://ecollection.ferc.gov/api/rssfeed"
    r = _http_get(url)
    root = ET.fromstring(r.content)
    items = []
    # RSS 2.0: channel > item
    for item in root.iter("item"):
        title = (item.findtext("title") or "").strip()
        if not HYDRO_KEYWORDS.search(title):
            continue
        link  = (item.findtext("link")  or "").strip()
        pub   = (item.findtext("pubDate") or "").strip()
        try:
            dt = datetime.datetime.strptime(pub[:25], "%a, %d %b %Y %H:%M:%S")
            date_iso = dt.date().isoformat()
        except ValueError:
            date_iso = ""
        items.append({
            "tag": "FERC", "tag_class": "ferc",
            "title": title, "date": date_iso,
            "source": "FERC eLibrary", "url": link
        })
        if len(items) >= max_items:
            break
    _log(f"  · FERC: {len(items)} hydropower-relevant items")
    return items


def fetch_caiso_notices(max_items: int = 5) -> List[Dict[str, str]]:
    """
    CAISO market notices — scrape the notices listing page.
    Source: https://www.caiso.com/library/notices-iso-news-and-information
    """
    from bs4 import BeautifulSoup
    url = "https://www.caiso.com/library/notices-iso-news-and-information"
    r = _http_get(url)
    soup = BeautifulSoup(r.text, "html.parser")
    items = []
    # CAISO notices page renders each notice as <article> or list rows;
    # be permissive: any anchor inside a date-prefixed row.
    for row in soup.select("article, li, tr"):
        a = row.find("a", href=True)
        if not a:
            continue
        title = a.get_text(strip=True)
        if not title or len(title) < 12:
            continue
        if not HYDRO_KEYWORDS.search(title):
            continue
        href = a["href"]
        if href.startswith("/"):
            href = "https://www.caiso.com" + href
        # Try to extract a date from the row text
        text = row.get_text(" ", strip=True)
        date_iso = ""
        m = re.search(r"(\d{1,2}/\d{1,2}/\d{4})", text)
        if m:
            try:
                date_iso = datetime.datetime.strptime(m.group(1), "%m/%d/%Y").date().isoformat()
            except ValueError:
                pass
        items.append({
            "tag": "CAISO", "tag_class": "market",
            "title": title, "date": date_iso,
            "source": "CAISO Notices", "url": href
        })
        if len(items) >= max_items:
            break
    _log(f"  · CAISO: {len(items)} hydropower-relevant items")
    return items


def fetch_bpa_news(max_items: int = 5) -> List[Dict[str, str]]:
    """
    BPA press releases — scrape the news releases listing page.
    Source: https://www.bpa.gov/about/newsroom/news-releases
    """
    from bs4 import BeautifulSoup
    url = "https://www.bpa.gov/about/newsroom/news-releases"
    r = _http_get(url)
    soup = BeautifulSoup(r.text, "html.parser")
    items = []
    for a in soup.find_all("a", href=True):
        title = a.get_text(strip=True)
        if not title or len(title) < 18:
            continue
        # BPA releases are prefixed "PR-NN-YY" — use as a filter signal too
        is_release = bool(re.match(r"PR-\d{2}-\d{2}", title))
        if not (is_release or HYDRO_KEYWORDS.search(title)):
            continue
        href = a["href"]
        if href.startswith("/"):
            href = "https://www.bpa.gov" + href
        # BPA listing pages don't always show dates inline; leave blank
        items.append({
            "tag": "BPA", "tag_class": "ops",
            "title": re.sub(r"^PR-\d{2}-\d{2}\s*", "", title),
            "date": "",
            "source": "BPA Newsroom", "url": href
        })
        if len(items) >= max_items:
            break
    _log(f"  · BPA: {len(items)} items")
    return items


def fetch_regulatory_pulse() -> List[Dict[str, str]]:
    """
    Pull from 3 feeds, dedupe by title, sort by date desc, cap at 6.
    """
    feeds = []
    for label, fn in [
        ("FERC",  fetch_ferc_filings),
        ("CAISO", fetch_caiso_notices),
        ("BPA",   fetch_bpa_news),
    ]:
        result = _safe(fn, label)
        if result:
            feeds.extend(result)

    # Dedupe by lowercase title
    seen = set()
    deduped = []
    for it in feeds:
        key = it["title"].lower().strip()[:80]
        if key in seen:
            continue
        seen.add(key)
        deduped.append(it)

    # Sort: items with dates first (newest), then dateless items
    deduped.sort(key=lambda x: x.get("date") or "0000", reverse=True)
    return deduped[:6]


# ══════════════════════════════════════════════════════════════════════
#  ARCHIVE (hand-maintained)
# ══════════════════════════════════════════════════════════════════════

def load_archive() -> List[Dict[str, str]]:
    if ARCHIVE_PATH.exists():
        try:
            return json.loads(ARCHIVE_PATH.read_text())
        except Exception as e:
            _log(f"  ⚠ archive.json parse error: {e}")
    return []


# ══════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════

def main() -> int:
    now = datetime.datetime.now(PT)
    _log(f"━━━ WECC Hydro fetch · {now.isoformat()} ━━━")

    # Basins
    _log("Basins:")
    basins: List[Dict[str, Any]] = []
    for b in BASINS:
        if   b["source"] == "snotel": data = _safe(lambda: fetch_snotel_basin(b["id"]), f"SNOTEL {b['id']}")
        elif b["source"] == "cdec":   data = _safe(lambda: fetch_cdec_sierra(b["id"]),  f"CDEC {b['id']}")
        elif b["source"] == "bc":     data = _safe(lambda: fetch_bc_snow(b["id"]),      f"BC {b['id']}")
        else: data = None
        if data:
            basins.append({"name": b["name"], **data})
            _log(f"  ✓ {b['name']}: {data.get('pct_median')}%")

    # Tier-1 markets
    _log("Markets:")
    bpa_mix = _safe(fetch_eia_bpa_mix,   "EIA-930 BPA")
    midc    = _safe(fetch_eia_midc_price, "EIA ICE bulk XLS Mid-C")
    wsf     = _safe(fetch_nwrfc_wsf,      "NWRFC WSF")

    # Ops
    _log("Ops:")
    usace_rows: List[Dict[str, Any]] = []
    for p in USACE_PROJECTS:
        data = _safe(lambda: fetch_usace_project(p["code"]), f"USACE {p['code']}")
        if data:
            usace_rows.append({"project": p["name"], **data})
    caiso = _safe(fetch_caiso_hydro, "CAISO OASIS")

    archive = load_archive()

    output = {
        "meta": {
            "last_updated_pt": now.strftime("%Y-%m-%d %H:%M PT"),
            "next_update_pt":  (now + datetime.timedelta(days=1)).strftime("%Y-%m-%d %H:%M PT"),
            "sources_ok": [k for k, v in {
                "bpa_mix": bpa_mix, "midc": midc, "wsf": wsf,
                "caiso": caiso, "usace": bool(usace_rows),
                "basins": bool(basins),
            }.items() if v],
        },
        "basins":  basins,
        "wsf":     wsf or {},
        "midc":    midc or {},
        "bpa_mix": bpa_mix or {},
        "usace":   usace_rows,
        "caiso":   caiso or {},
        "archive": archive,
    }

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(json.dumps(output, indent=2, default=str))
    _log(f"━━━ wrote {OUT_PATH} ({OUT_PATH.stat().st_size} bytes) ━━━")
    return 0


if __name__ == "__main__":
    sys.exit(main())
